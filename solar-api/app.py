#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
app.py — 태양광 계약서류 자동입력 Flask API 서버
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Word 템플릿 방식으로 서류 생성 → ZIP 반환
Render.com / Railway 에 그대로 배포 가능

엔드포인트:
  GET  /api/health              서버 상태 확인
  POST /api/generate            단건 생성 → ZIP
  POST /api/generate_batch      엑셀 업로드 일괄 → ZIP
  GET  /api/download_template   엑셀 양식 다운로드
  GET  /api/check_templates     템플릿 파일 상태 확인
"""

import io
import os
import re
import json
import zipfile
import tempfile
import logging
from datetime import datetime
from pathlib import Path

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

# ── 로깅 설정 ─────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger(__name__)

# ── 경로 설정 ─────────────────────────────────────────
BASE_DIR = Path(__file__).parent
TEMPLATE_DIR = BASE_DIR / "templates"
FONT_PATH = BASE_DIR / "NanumGothic.ttf"

# ── docx_engine 임포트 ────────────────────────────────
import sys
sys.path.insert(0, str(BASE_DIR))

try:
    from docx_engine import generate_all_docs, check_templates, _normalize_data
    log.info("✅ docx_engine 로드 성공")
except ImportError as e:
    log.error(f"❌ docx_engine 로드 실패: {e}")
    generate_all_docs = None

# ── 엑셀 파싱 (openpyxl) ─────────────────────────────
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log.warning("openpyxl 없음 — 엑셀 기능 비활성")

# ── Flask 앱 ──────────────────────────────────────────
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20MB


# ════════════════════════════════════════════════════════
#  헬스체크
# ════════════════════════════════════════════════════════

@app.route("/api/health")
def health():
    tpl_status = check_templates() if generate_all_docs else {}
    return jsonify({
        "status": "ok",
        "version": "2.0",
        "engine": "docx_word_template",
        "templates": tpl_status,
        "excel_support": EXCEL_AVAILABLE,
        "timestamp": datetime.now().isoformat()
    })


# ════════════════════════════════════════════════════════
#  템플릿 상태 확인
# ════════════════════════════════════════════════════════

@app.route("/api/check_templates")
def api_check_templates():
    if not generate_all_docs:
        return jsonify({"error": "docx_engine 로드 실패"}), 500
    return jsonify(check_templates())


# ════════════════════════════════════════════════════════
#  단건 생성
# ════════════════════════════════════════════════════════

@app.route("/api/generate", methods=["POST"])
def api_generate():
    if not generate_all_docs:
        return jsonify({"error": "서류 생성 엔진 초기화 실패"}), 500

    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({"error": "JSON 데이터가 없습니다."}), 400

    # 필수 항목 검증
    required = ["상호명", "대표자명"]
    missing = [f for f in required if not data.get(f, "").strip()]
    if missing:
        return jsonify({"error": f"필수 항목 누락: {', '.join(missing)}"}), 400

    log.info(f"📄 단건 생성 요청: {data.get('상호명')}")

    with tempfile.TemporaryDirectory() as tmp:
        success, errors = generate_all_docs(data, tmp)

        if not success:
            log.error(f"생성 실패: {errors}")
            return jsonify({"error": "서류 생성 실패", "details": errors}), 500

        # ZIP 묶기
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for filepath in success:
                zf.write(filepath, Path(filepath).name)
        buf.seek(0)

    safe = re.sub(r'[\\/:*?"<>|]', "_", data.get("상호명", "발전소"))
    fname = f"태양광계약서류_{safe}_{datetime.now().strftime('%Y%m%d')}.zip"

    log.info(f"✅ 생성 완료: {len(success)}개 → {fname}")

    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=fname,
    )


# ════════════════════════════════════════════════════════
#  엑셀 업로드 → 파싱 (미리보기용)
# ════════════════════════════════════════════════════════

@app.route("/api/upload_excel", methods=["POST"])
def api_upload_excel():
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "openpyxl 미설치"}), 500
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "파일명이 비어있습니다."}), 400

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        f.save(tmp.name)
        try:
            rows = _load_excel(tmp.name)
        except Exception as e:
            return jsonify({"error": str(e)}), 400
        finally:
            os.unlink(tmp.name)

    return jsonify({"rows": rows, "count": len(rows)})


# ════════════════════════════════════════════════════════
#  엑셀 일괄 생성
# ════════════════════════════════════════════════════════

@app.route("/api/generate_batch", methods=["POST"])
def api_generate_batch():
    if not generate_all_docs:
        return jsonify({"error": "서류 생성 엔진 초기화 실패"}), 500
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "openpyxl 미설치"}), 500
    if "file" not in request.files:
        return jsonify({"error": "엑셀 파일이 없습니다."}), 400

    excel_file = request.files["file"]
    selected_str = request.form.get("selected", "")

    try:
        selected_indices = [int(i) for i in selected_str.split(",")
                            if i.strip().isdigit()]
    except Exception:
        selected_indices = []

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        excel_file.save(tmp.name)
        try:
            rows = _load_excel(tmp.name)
        except Exception as e:
            return jsonify({"error": str(e)}), 400
        finally:
            os.unlink(tmp.name)

    if not rows:
        return jsonify({"error": "엑셀에 데이터가 없습니다."}), 400

    if selected_indices:
        rows = [rows[i] for i in selected_indices if 0 <= i < len(rows)]

    log.info(f"📦 일괄 생성 요청: {len(rows)}건")

    buf = io.BytesIO()
    all_errors = []

    with tempfile.TemporaryDirectory() as batch_tmp:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for row in rows:
                safe = re.sub(r'[\\/:*?"<>|]', "_", row.get("상호명", "발전소"))
                sub_dir = os.path.join(batch_tmp, safe)
                success, errors = generate_all_docs(row, sub_dir)
                all_errors.extend(errors)
                for fp in success:
                    zf.write(fp, f"{safe}/{Path(fp).name}")

    buf.seek(0)
    fname = f"태양광계약서류_일괄_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

    log.info(f"✅ 일괄 완료: {len(rows)}건, 오류: {len(all_errors)}건")

    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=fname,
    )


# ════════════════════════════════════════════════════════
#  엑셀 양식 다운로드
# ════════════════════════════════════════════════════════

@app.route("/api/download_template")
def api_download_template():
    # 기존 파일 있으면 바로 제공
    tpl = BASE_DIR / "발전소목록_입력양식.xlsx"
    if tpl.exists():
        return send_file(str(tpl), as_attachment=True,
                         download_name="발전소목록_입력양식.xlsx")

    # 없으면 동적 생성
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "openpyxl 미설치"}), 500

    try:
        buf = _create_excel_template()
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="발전소목록_입력양식.xlsx",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ════════════════════════════════════════════════════════
#  내부 헬퍼
# ════════════════════════════════════════════════════════

HEADER_ALIASES = {
    "상호명":           ["상호명", "회사명", "사업장명", "법인명"],
    "대표자명":         ["대표자명", "대표자", "대표"],
    "사업자등록번호":   ["사업자등록번호", "사업자번호", "사업자No"],
    "법인등록번호":     ["법인등록번호", "법인번호"],
    "사업장주소":       ["사업장주소", "주소", "사업장 주소"],
    "발전기주소":       ["발전기주소", "발전기 주소", "발전기 설치주소",
                        "발전기 설치주소 ★", "설치주소", "태양광주소"],
    "설비용량":         ["설비용량", "용량(kW)", "용량", "kW",
                        "설비용량(kW)", "설비 용량"],
    "전기사업자등록번호": ["전기사업자등록번호", "전기사업자번호"],
    "연락처":           ["연락처", "전화번호", "전화", "연락전화", "담당자전화"],
    "이메일":           ["이메일", "email", "Email", "E-mail", "메일"],
    "예금주명":         ["예금주명", "예금주"],
    "은행명":           ["은행명", "은행"],
    "계좌번호":         ["계좌번호", "계좌"],
}


def _load_excel(path: str) -> list:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # 헤더 행 찾기
    header_row_idx = None
    col_map = {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        for ci, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            for key, aliases in HEADER_ALIASES.items():
                if cell_str in aliases and key not in col_map:
                    col_map[key] = ci
                    header_row_idx = ri
    wb.close()

    if header_row_idx is None:
        raise ValueError("헤더 행을 찾을 수 없습니다. 양식을 확인해주세요.")

    # 데이터 읽기
    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2.active
    rows = []

    for ri, row in enumerate(ws2.iter_rows(min_row=header_row_idx + 2,
                                            values_only=True)):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        item = {}
        for key, ci in col_map.items():
            if ci < len(row) and row[ci] is not None:
                item[key] = str(row[ci]).strip()
        if item.get("상호명") or item.get("대표자명"):
            rows.append(item)

    wb2.close()
    return rows


def _create_excel_template() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발전소목록"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    headers = list(HEADER_ALIASES.keys())
    required = {"상호명", "대표자명", "사업자등록번호", "사업장주소", "설비용량", "연락처"}

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h + (" ★" if h in required else ""))
        cell.font = Font(bold=True, size=10,
                         color="FFFFFF" if h in required else "333333")
        cell.fill = PatternFill("solid",
                                fgColor="1E40AF" if h in required else "64748B")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = max(len(h) * 2.2, 14)

    # 예시 데이터
    example = {
        "상호명": "홍길동태양광발전(주)",
        "대표자명": "홍길동",
        "사업자등록번호": "123-45-67890",
        "법인등록번호": "110111-1234567",
        "사업장주소": "서울특별시 강남구 테헤란로 123",
        "발전기주소": "전라남도 영암군 삼호읍 나불리 100",
        "설비용량": "99.9",
        "전기사업자등록번호": "2023-서울-0001",
        "연락처": "010-1234-5678",
        "이메일": "test@example.com",
        "예금주명": "홍길동태양광발전",
        "은행명": "국민은행",
        "계좌번호": "123456-78-901234",
    }
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=example.get(h, ""))
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════
#  메인
# ════════════════════════════════════════════════════════

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    debug = os.environ.get("DEBUG", "false").lower() == "true"

    log.info("=" * 55)
    log.info("  ☀  태양광 계약서류 자동입력 API 서버 v2.0")
    log.info(f"  포트: {port}")
    log.info(f"  엔진: Word 템플릿 방식")
    log.info(f"  템플릿: {TEMPLATE_DIR}")
    log.info("=" * 55)

    # 템플릿 상태 출력
    if generate_all_docs:
        status = check_templates()
        for name, exists in status.items():
            icon = "✅" if exists else "❌"
            log.info(f"  {icon} {name}")

    app.run(host="0.0.0.0", port=port, debug=debug)
